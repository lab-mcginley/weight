from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill
from datetime import date as Date
import os
from calendar import Calendar
import serial
import time
import logging


def date_working(today) ->str:
    check_date = 0
    while check_date == 0:
        print(f"\n Today is {today.isoformat()}")
        date = input("\n Press ENTER if you are submitting for today, or input date in yyyy-mm-dd format: ")

        if date == '':
            date = today.isoformat()
            check_date = 1
        else:
            try:
                calendar = [i for i in get_calendar(check_lastday(date))]
                if date not in calendar:
                    print('\nNo such day in the calendar\n\n\n')
                    check_date = 0
                else:
                    check_date = 1
            
            except Exception as e:
                print(e)
                print('\nInput date not valid\n\n\n')
                check_date = 0
    
    return date


def check_lastday(date) ->int:
    """
    Special treatment to date in first or last week

    Parameters
    ----------
    date : str
        date in 'yyyy-mm-dd' format

    Returns
    -------
    int
        working year.

    """
    c = Calendar()
    year = int(date.split('-')[0])
    first_week = [i.isoformat() for i in c.yeardatescalendar(year, width=12)[0][0][0]]
    last_week = [i.isoformat() for i in c.yeardatescalendar(year, width=12)[0][-1][-1]]
    
    #if 12/31 is Sun, move to next year
    if date[5:] == '12-31' and date == last_week[-1]:
        return year+1    
    #if 1/1 not Sun, move to previous year
    elif date in first_week and f'{year-1}-12-30' in first_week:
        return year-1
    else:
        return year


def get_calendar(year):
    """
    Generate yearly calender from first Sunday to last Saturday

    Returns
    -------
    list
        list of datetime.date object.

    """
    c = Calendar()
    yearly = []
    for month in range(1,12,1):
        day_list = list(c.itermonthdates(year,month))
        day_zero = list(c.itermonthdays(year,month))
        dates = [dd.isoformat() for ii,dd in enumerate(day_list) if day_zero[ii]!=0]
        yearly.append(dates)
    
    #December list is to get full week calender if last week containes dates in Janurary
    #But November in the first week need to be remove or it will be duplicated   
    dec_list = list(c.itermonthdates(year,12))
    yearly.append([i.isoformat() for i in dec_list if i.month != 11])
    
    yearly = [daily for monthly in yearly for daily in monthly]
    
    jan_list = list(c.itermonthdates(year,1))
    if jan_list[0].isoformat()[5:] == '01-01':
        end = c.monthdatescalendar(year, 12)[-1][5].isoformat()
        
        start_i = 0
        yearly.insert(0, f'{year-1}-12-31')
        
        end_i = [i for i,a in enumerate(yearly) if a == end][0]
    else:
        #start = First Sunday after 1/1
        #end = Last Saturday of the year
        start = c.monthdatescalendar(year, 1)[0][6].isoformat()
        end = c.monthdatescalendar(year, 12)[-1][5].isoformat()
    
        start_i = [i for i,a in enumerate(yearly) if a == start][0]
        end_i = [i for i,a in enumerate(yearly) if a == end][0]
    
    return yearly[start_i:end_i+1]


def get_working_row(date):
    """
    Return row index for working date

    Parameters
    ----------
    date : str
        date in yyyy-mm-dd format.

    Returns
    -------
    int
        row index of working date

    """
    year = check_lastday(date)
    return [i for i,a in enumerate(get_calendar(year), start=8) if a == date][0]


def get_deprive_row(date, sheet) ->int:
    """
    Return deprive date index cloest to working date

    Parameters
    ----------
    date : str
        date in yyyy-mm-dd format
    sheet : worksheet
        current working openpyxl worksheet

    Returns
    -------
    int
        row index of deprive date

    """
    year = check_lastday(date)
    i_current = [i for i,a in enumerate(get_calendar(year), start=8) if a == date][0]
    i_deprive = [i_current-i for i in range(8) if sheet.cell(row=i_current-i, column=3).value == 'D']
    
    if i_deprive != []:
        return i_deprive[0]
    else:
        print("No Deprivation Record Within 7 Days")
        return i_current+1


def new_sheet(mouseID, year):
    wb.create_sheet(mouseID)
    
    sheet = wb[mouseID]
    sheet.append(['Date', 'Days', 'Deprive', 'Mass (g)', '',
                  '% of weight on deprived day', 'FEED (g)'])
    sheet.move_range('A1:G1', rows=6)
    calender = get_calendar(year)
    #Calerdar
    for i,a in enumerate(calender, start=8):
        sheet.cell(row=i, column=1).value = a
        sheet.cell(row=i, column=1).alignment = Alignment(horizontal='center')
    
    #Week Days
    _week = ['SUN', 'MON', 'TUE', 'WED', 'THR', 'FRI', 'SAT']
    for i in range(8,len(calender)+8):
        current_cell = sheet.cell(row=i, column=2)
        current_cell.value = _week[(i-8)%7]
        if _week[(i-8)%7] == 'SUN':
            current_cell.fill = fill_red()
        current_cell.alignment = Alignment(horizontal='center')
    
    #Default deprivation every Sunday
    for i in range(8,len(calender)+8,7):
        sheet.cell(row=i, column=3).value = 'D'
        sheet.cell(row=i, column=3).alignment = Alignment(horizontal='center')
                    
    sheet['A1'] = 'ID'
    sheet['B1'] = mouseID
    """Width nunber*10 = Pixel in excel"""
    sheet.column_dimensions['A'].width = 13
    sheet.column_dimensions['B'].width = 8
    sheet.column_dimensions['F'].width = 27  
    
    return sheet

def get_weight(manual):
    if not manual:
        timeout = time.time()
        weighting=1
        ser = serial.Serial(port='COM3', timeout=1, xonxoff=True)
        
        while weighting:
            print('Waiting reading from the scale, please press print.')
            time.sleep(2)
            reading = ser.readline()
            if reading:
                weight = float(str(reading).split('\\')[0][2:])
                overtime=0
                weighting=0
                ser.close()
            elif time.time()-timeout > 30:
                overtime=1
                weighting=0
                weight=0
                ser.close()
                              
        if 100 > weight > 1:
            #print(f'Mouse weight: {weight} g')
            manual=0
            return float(weight), manual
        
        elif overtime:
            manual = int(input("Not getting weight from the scale, do you wish to manually enter mouse weight? (1:YES / 0:NO): "))
            
            while manual != 0 and manual != 1:
                manual = int(input("please only enter 0 or 1: "))
            
            if manual:
                weight = input("Please input mouse weight in g: ")
                close_ser(ser)
                return float(weight), manual
            else:
                close_ser(ser)
                weight, manual = get_weight(manual)               
                return float(weight), manual
        else:
            print("Abnormal weight from scale")
            return float(weight), manual
        
    elif manual:
        weight = input("Please input mouse wight in g: ")
        if 'ser' in locals():
            close_ser(ser)
        return float(weight), manual
    
def feed_amout(ratio):
    if ratio >= 0.9:
        feed = 0.8
    elif ratio >= 0.88:
        feed = 1.0
    elif ratio >= 0.86:
        feed = 1.2
    elif ratio >= 0.84:
        feed = 1.6
    elif ratio >= 0.82:
        feed = 1.9
    elif ratio >= 0.80:
        feed = 2.5
    else:
        feed = 3.0
        
    return feed

def record_feeding(row_working, ratio, feed):
    sheet.cell(row=row_working, column=6).value = float('%.2f'%(ratio*100))
    sheet.cell(row=row_working, column=6).alignment = Alignment(horizontal='center')
    sheet.cell(row=row_working, column=7).value = float('%.1f'%feed)
    wb.save(file)
    
def fill_red():
    fill = PatternFill(start_color='80FF0000',
           end_color='80FF0000',
           fill_type='solid')
    
    return fill

def close_ser(ser):
    if ser.isOpen():
        ser.close()
    del ser


if __name__ == "__main__":
    console = logging.StreamHandler()
    formatter = logging.Formatter('%(message)s')
    console.setFormatter(formatter)
    handlers = [logging.FileHandler('logs.log'), console]
    logging.basicConfig(level=logging.INFO, format='%(asctime)s-> %(message)s', datefmt='%Y/%m/%d-%H:%M', handlers=handlers)
    
    today = Date.today()
    date = date_working(today)
    year = check_lastday(date)
    
    file = f'weight_{year}.xlsx'
    
    if os.path.isfile(file):
        wb = load_workbook(file)
    else:
        print(f"\n{file} does not exist, creating new entry.\n")
        wb = Workbook()
        new_sheet('template', year)
    
    logging.info(f"Entering for {date}, save with {file}")
    
    print('~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~')
    print('If wish to manualy key in mouse weight, entering: MANUAL')
    print('When finished, save by entering: STOP')
    print(f'**Be sure to close {file} before recording mouse weight**')
    print('~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~')
    
    manual = 0 
    mouseID = ()
    while mouseID != 'STOP':
        mouseID = input("Please Enter Mouse ID: ")
        if mouseID == '':
            print('No ID entered')
            continue
        mouseID = mouseID.upper()
        _entering = 0
        
        if mouseID == 'STOP':       
            break       
        elif mouseID == 'MANUAL':
            manual = 1        
        elif mouseID in wb.sheetnames:
            sheet = wb[mouseID]
            _entering = 1
        else:
            print("Mouse with this ID doesn't exist, create new sheet for it?")
            ans = input("YES=1 / NO=0: ")
            while ans != '0' and ans != '1':
                ans = input("Please only enter 0 or 1: ")
            
            if int(ans):
                sheet = new_sheet(mouseID, year)
                _entering = 1
            else:
                _entering = 0
    
        if _entering:
            row_working = get_working_row(date)
            row_deprive = get_deprive_row(date, sheet)
            
            #no deprived date detected
            if row_deprive == row_working+1:
                break
            #input for deprived date
            elif row_deprive == row_working:
                working_weight, manual = get_weight(manual)
                if working_weight == None or 1 > working_weight > 100:
                    logging.warning(f"WARNING: Abnormal derpive Weight at {working_weight} g")       
                else:
                    sheet.cell(row=row_working, column=4).value = working_weight
                    logging.info(f"ID: {mouseID}, Deprive Weight: {working_weight} g")
            #input for traning date
            else:
                try:
                    weight_deprive = float(sheet.cell(row=row_deprive, column=4).value)
                except:
                    logging.warning("WARNING: No deprive weight detected")
                    add_deprive = input("Please enter deprive weight or enter STOP to abort: ")
                    
                    if add_deprive.upper() == 'STOP':
                        break
                    else:
                        try:
                            weight_deprive = float(add_deprive)
                            logging.info(f"Using {weight_deprive} g as deprive weight for {mouseID}, please manually add it to the excel file later")
                            time.sleep(2)
                        except:
                            print("WARNING: Not getting the right number")
                            break
                    
                if weight_deprive == None or 1 > weight_deprive > 100:
                    logging.warning(f"WARNING: Abnormal derpive Weight at {weight_deprive}")                 
                else:
                    weight_working, manual= get_weight(manual)
                    
                    if weight_working == None or 1 >= weight_working >= 100:
                        logging.warning(f"WARNING: Abnormal weight at {weight_working} g")
                    else:
                        sheet.cell(row=row_working, column=4).value = weight_working
                        ratio = weight_working/weight_deprive                  
                        feed = feed_amout(ratio)
                        logging.info(f'ID: {mouseID}, Weight: {weight_working} g')
                        logging.info(f"Please feed {feed} g.\n")
                        record_feeding(row_working, ratio, feed)
                        
                                   
    logging.shutdown()
    wb.save(file)


